'''
MEGA ULTRA SOPA DE LETRAS GENERATOR MADE BY DIOGO VALA
'''


import random
import string
from openpyxl import load_workbook
import pandas as pd
from win32com import client
import win32api
import pathlib
import sys
import time

class game_c:
    def __init(self):
        self.words=list()
        self.grid=[['_' for _ in range(grid_size)] for _ in range(grid_size)]

def readWords():

    words_short=open('words_short.txt', encoding="utf8").readlines()
    words_med=open('words_med.txt', encoding="utf8").readlines()
    words_long=open('words_long.txt', encoding="utf8").readlines()

    words_short = [random.choice(words_short).upper().strip() for _ in range(int(n_words/3))]
    words_short.sort()
    words_short.sort(key=len)

    words_med = [random.choice(words_med).upper().strip() for _ in range(int(n_words/3))]
    words_med.sort()
    words_med.sort(key=len)

    words_long = [random.choice(words_long).upper().strip() for _ in range(int(n_words/3))]
    words_long.sort()
    words_long.sort(key=len)

    words=words_short+words_med+words_long

    return words

def generateSoup(words):

    game = game_c()
    game.grid = [['_' for _ in range(grid_size)] for _ in range(grid_size)]
    game.words=list()
    orientations = ['leftright', 'rightleft', 'updown', 'downup', 'diagonalup', 'diagonaldown', 'invdiagonalup', 'invdiagonaldown']

    for word in reversed(words):
        word_length = len(word)

        max_it=0

        placed = False
        while not placed:
            max_it=max_it+1
            if(max_it > 20000):
                print("fail")
                word=''
                break
            orientation = random.choice(orientations)
            if orientation== 'leftright':
                step_x=1
                step_y=0
            elif orientation == 'rightleft':
                step_x=-1
                step_y=0
            elif orientation == 'updown':
                step_x=0
                step_y=1
            elif orientation == 'downup':
                step_x=0
                step_y=-1
            elif orientation == 'diagonalup':
                step_x=1
                step_y=1
            elif orientation == 'diagonaldown':
                step_x=1
                step_y=-1
            elif orientation == 'invdiagonalup':
                step_x=-1
                step_y=1
            elif orientation == 'invdiagonaldown':
                step_x=-1
                step_y=-1


            x_position = random.randint(0,grid_size-1)
            y_position = random.randint(0,grid_size-1)

            ending_x = x_position + word_length*step_x
            ending_y = y_position + word_length*step_y

            if ending_x >= grid_size or ending_x < 0: continue
            if ending_y >= grid_size or ending_y < 0: continue

            failed = False

            for i in range(word_length):
                character = word[i]

                new_position_x = x_position + i*step_x
                new_position_y = y_position + i*step_y
                character_at_new_position = game.grid[new_position_x][new_position_y]
                if character_at_new_position == '_' or character_at_new_position == character:
                    continue
                else:
                    failed = True
                    break

            if failed:
                continue
            else:
                for i in range(word_length):
                    character= word[i]
                    new_position_x = x_position + i * step_x
                    new_position_y = y_position + i * step_y
                    game.grid[new_position_x][new_position_y] = character
                placed=True
        game.words.append(word)        

    valid_chars=list('ABCDEFGHIJLMNOPQRSTUVXZ')
    for x in range(grid_size):
        for y in range(grid_size):
            if game.grid[x][y]=='_':
                game.grid[x][y]=random.choice(valid_chars)
                #game.grid[x][y]='_' #debug
    game.words=reversed(game.words)
    return game

def exportExcel(game1, words1, game2, words2):

    #Load file
    wb = load_workbook(filename)
    ws = wb['Sheet2']

    #Game1 grid
    list=[]
    for x in range(grid_size):
        line=''
        for y in range(grid_size):
            line += ' '.join(game1[x][y])
        list.append(line)
    df_new = pd.DataFrame({'Game1': list})

    for index, row in df_new.iterrows():
        cell = 'A%d' % (index + 1)
        ws[cell] = row[0]

    #Game1 words
    list=[]
    for word in words1:
        list.append(word)
    df_new = pd.DataFrame({'Words1': list})

    for index, row in df_new.iterrows():
        cell = 'B%d' % (index + 1)
        ws[cell] = row[0]

    #Game2 grid
    list=[]
    for x in range(grid_size):
        line=''
        for y in range(grid_size):
            line += ' '.join(game2[x][y])
        list.append(line)
    df_new = pd.DataFrame({'Game2': list})

    for index, row in df_new.iterrows():
        cell = 'C%d' % (index + 1)
        ws[cell] = row[0]

    #Game2 words
    list=[]
    for word in words2:
        list.append(word)
    df_new = pd.DataFrame({'Words2': list})

    for index, row in df_new.iterrows():
        cell = 'D%d' % (index + 1)
        ws[cell] = row[0]

    wb.save(filename)

def convertPDF(n, n_files):
    excel_file = "sopa.xlsx"
    pdf_file = "sopa"+str(n)+".pdf"
    excel_path = str(pathlib.Path.cwd() / excel_file)
    pdf_path = str(pathlib.Path.cwd() / pdf_file)
    excel = client.DispatchEx("Excel.Application")
    excel.Visible = 0
    wb = excel.Workbooks.Open(excel_path)
    ws = wb.Worksheets[1]
    try:
        wb.SaveAs(pdf_path, FileFormat=57)
    except Exception as e:
        print("Failed to convert")
        print(str(e))
    finally:
        wb.Close()
        excel.Quit()

def pdfmerge(n_files):
    from PyPDF2 import PdfFileMerger
    import os
    pdfs=[]
    for n in range(n_files):
        pdfs.append('sopa'+str(n)+'.pdf')

    merger = PdfFileMerger()

    for pdf in pdfs:
        merger.append(pdf)

    merger.write("Livro.pdf")
    merger.close()

    for pdf in pdfs:
        os.remove(pdf)

def main():
    n_files = int(sys.argv[1])
    maintic=time.perf_counter()
    game1 = game_c()
    game2 = game_c()
    for n in range(n_files):
        print(f"Game {n}")
        tic=time.perf_counter()
        game1.words = readWords()
        game2.words = readWords()
        toc=time.perf_counter()
        print(f"Words read in: {toc-tic:0.3f} seconds")

        game1 = generateSoup(game1.words)
        game2 = generateSoup(game2.words)

        tic=time.perf_counter()
        exportExcel(game1.grid, game1.words, game2.grid, game2.words)
        toc=time.perf_counter()
        print(f"Excel generated in: {toc-tic:0.3f} seconds")

        tic=time.perf_counter()
        convertPDF(n, n_files)
        toc=time.perf_counter()
        print(f"PDF generated in: {toc-tic:1.3f} seconds")
        print("\n")

  
    pdfmerge(n_files)

    maintoc=time.perf_counter()
    print(f"{n_files} game pages generated in: {maintoc-maintic:.3f} seconds")

filename='sopa.xlsx'
n_words = 24 #Words in a game
grid_size = 20 #Size of game
main()



